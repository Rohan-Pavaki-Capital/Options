"""Raw balance-sheet Excel export — same Stage 1 (locate) + Stage 2 (parse)
as the standardize pipeline, but NO Stage 3/4: instead of mapping into the
Damodaran schema, every line item and every value from the parsed
balance-sheet page(s) is written to an Excel workbook exactly as printed —
no summarizing, no standardizing, no LLM call.
"""

import logging
import os
import re

from ..preprocessing import parser, pdf_locator

logger = logging.getLogger("balance_sheet.raw_excel")

# Dash cells mean "no value printed" — kept as text so the sheet mirrors the
# filing (converting them to 0 would be an interpretation, not a copy).
_DASH_CELLS = {"-", "--", "—", "–"}
_NUM_RE = re.compile(r"\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?")
_SEPARATOR_ROW_RE = re.compile(r"^\|[\s:\-|]+\|$")


def _coerce_cell(raw: str):
    """Return the cell as a number when it is purely a printed figure
    ($ sign, digit-group commas, parenthesized negatives), otherwise the
    original text unchanged. Values are copied exactly — never recomputed."""
    text = raw.strip().strip("*").strip()
    if not text or text in _DASH_CELLS:
        return text
    candidate = text.replace("$", "").strip()
    negative = candidate.startswith("(") and candidate.endswith(")")
    if negative:
        candidate = candidate[1:-1].strip()
    if _NUM_RE.fullmatch(candidate):
        value = float(candidate.replace(",", ""))
        if negative:
            value = -value
        return int(value) if value == int(value) else value
    return text


def markdown_to_rows(markdown: str) -> list[list]:
    """Convert LlamaParse markdown to spreadsheet rows, preserving every
    line as-is: table rows become one cell per column, plain text lines
    (statement title, company name, "(in millions)" …) go in column A.
    Only markdown table-separator rows (|---|---|) are dropped."""
    rows: list[list] = []
    for line in markdown.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("|"):
            if _SEPARATOR_ROW_RE.fullmatch(stripped):
                continue  # markdown |---| separator, not filing content
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            rows.append([_coerce_cell(c) for c in cells])
        else:
            rows.append([stripped.lstrip("#").strip()])
    return rows


def build_raw_workbook(markdown: str, output_path: str,
                       source_pages: list | None = None) -> str:
    """Write the as-printed balance sheet to an .xlsx at output_path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    rows = markdown_to_rows(markdown)
    if not rows:
        raise RuntimeError("Parsed markdown contained no balance-sheet rows.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    bold = Font(bold=True)
    for row in rows:
        ws.append(row)
        if len(row) == 1 and isinstance(row[0], str):
            ws.cell(row=ws.max_row, column=1).font = bold  # heading lines
    if source_pages:
        ws.append([])
        ws.append([f"Source: filing page(s) {', '.join(str(p) for p in source_pages)}"])

    # Readable widths: wide label column, regular value columns.
    max_cols = max(len(r) for r in rows)
    ws.column_dimensions["A"].width = 60
    for idx in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    wb.save(output_path)
    logger.info("Raw balance-sheet workbook written to %s", output_path)
    return output_path


def run_raw_pipeline(pdf_path: str, output_path: str) -> dict:
    """Stage 1 (locate) + Stage 2 (parse) from the standardize pipeline, then
    the as-is Excel export. Raises on failure (the caller maps it to HTTP)."""
    located = pdf_locator.locate_balance_sheet(pdf_path)
    temp_pdf = located["temp_pdf_path"]
    source_pages = located["page_numbers"]
    logger.info("Captured pages %s from %s", source_pages, pdf_path)
    try:
        markdown = parser.parse_to_markdown(temp_pdf)
    finally:
        if temp_pdf and os.path.exists(temp_pdf):
            try:
                os.remove(temp_pdf)
            except OSError:
                logger.warning("Could not delete temp PDF %s", temp_pdf)

    build_raw_workbook(markdown, output_path, source_pages)
    return {
        "excel_path": output_path,
        "source_pages": source_pages,
        "warnings": located.get("warnings", []),
    }
